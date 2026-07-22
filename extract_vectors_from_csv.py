import numpy as np
import pandas as pd
import openpyxl
import matplotlib
matplotlib.use('Agg') # Use Agg backend for non-GUI environments
import matplotlib.pyplot as plt
from matplotlib.animation import FuncAnimation # Retained for potential future use or if animation logic is re-added
import io
import os
from natsort import natsorted
from scipy.spatial import ConvexHull
from scipy.signal import find_peaks
from scipy.stats import iqr # For outlier detection
import warnings # To suppress specific warnings
import logging
import seaborn as sns
import re
from scipy import signal
import datetime # Import datetime for proper handling
import pywt

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

# Suppress specific openpyxl warning about data validation (if still relevant, keep)
warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')
warnings.filterwarnings("ignore", message="'S' is deprecated", category=FutureWarning)
warnings.filterwarnings("ignore", message="'M' is deprecated", category=FutureWarning)

def analyze_circadian_rhythm(csv_dir):
    """Analyze circadian rhythms using continuous wavelet transform."""
    try:
        feature_files = [f for f in os.listdir(csv_dir) 
                        if f.startswith('feature_') and f.endswith('.csv')]
        
        if not feature_files:
            print("No feature files found in directory")
            return None, None, None
            
        # Create hourly activity dataset
        all_data = []
        for f in feature_files:
            df = pd.read_csv(os.path.join(csv_dir, f), header=None, 
                            names=['x', 'y', 'datetime', 'rep'])
            df['datetime'] = pd.to_datetime(df['datetime'])
            df['displacement'] = np.sqrt(
                np.diff(df['x'], append=df['x'].iloc[-1])**2 + 
                np.diff(df['y'], append=df['y'].iloc[-1])**2
            )
            all_data.append(df)
        
        # Combine and resample to hourly activity
        combined_df = pd.concat(all_data)
        hourly_activity = combined_df.set_index('datetime').resample('h').agg({
            'displacement': 'mean'
        }).fillna(0)
        
        # Prepare data for wavelet transform
        data = hourly_activity['displacement'].values
        N = len(data)
        
        if N < 24:  # Need at least 24 hours of data
            print("Not enough data points for analysis")
            return None, None, None
            
        # Perform continuous wavelet transform
        scales = np.arange(1, 49)  # Look for patterns up to 48 hours
        coef, freqs = pywt.cwt(data, scales, 'morl')  # Using Morlet wavelet
        
        return hourly_activity, coef, scales
        
    except Exception as e:
        print(f"Error in circadian analysis: {e}")
        return None, None, None

def plot_circadian_analysis(hourly_activity, coef, scales, plot_limits=None):
    """Plot circadian rhythm analysis results with configurable limits."""
    try:
        if hourly_activity is None or coef is None or scales is None:
            print("No valid data for circadian analysis plotting")
            return None
        
        plt.style.use('ggplot') # Apply science and grid styles
        # Main analysis figure
        fig = plt.figure(figsize=(8, 6))
        
        # Activity over time
        ax1 = plt.subplot(211)
        ax1.plot(hourly_activity.index, hourly_activity['displacement'])
        ax1.set_title('Plant Movement Activity Over Time')
        ax1.set_xlabel('Time')
        ax1.set_ylabel('Movement Magnitude')
        
        if plot_limits and 'activity' in plot_limits:
            ax1.set_ylim(plot_limits['activity']['ymin'], 
                    plot_limits['activity']['ymax'])
        plt.xticks(rotation=45)
        
        # Wavelet transform scalogram
        ax2 = plt.subplot(212)
        times = np.arange(len(hourly_activity))
        
        vmin = 0
        vmax = 12
        if plot_limits and 'scalogram' in plot_limits:
            vmin = plot_limits['scalogram']['vmin']
            vmax = plot_limits['scalogram']['vmax']
            
        im = ax2.pcolormesh(times, scales, np.abs(coef), 
                        cmap='viridis', 
                        shading='gouraud',
                        vmin=vmin,
                        vmax=vmax)
        
        ax2.set_title('Wavelet Transform Scalogram')
        ax2.set_xlabel('Time (hours)')
        ax2.set_ylabel('Period (hours)')
        ax2.set_ylim([1, 48])
        
        cbar = plt.colorbar(im, ax=ax2)
        cbar.set_label('Power (Fixed Scale)')
        
        ax2.axhline(y=24, color='r', linestyle='--', alpha=0.5)
        ax2.text(len(times)-1, 24, '24h', color='r', alpha=0.5)
        
        plt.tight_layout()
        return fig
            
    except Exception as e:
        print(f"Error plotting circadian analysis: {e}")
        return None, None

def create_whole_plant_visualization(csv_dir, plot_limits=None):
    """Create separate comprehensive visualizations of entire plant motion."""
    try:
        feature_files = [f for f in os.listdir(csv_dir) 
                        if f.startswith('feature_') and f.endswith('.csv')]
        
        figures = {}  # Dictionary to store all figures
        
        # 1. Combined Feature Paths Plot
        fig_paths = plt.figure(figsize=(10, 8))
        ax_paths = fig_paths.add_subplot(111)
        colors = plt.cm.rainbow(np.linspace(0, 1, len(feature_files)))
        
        all_data = []  # Store all feature data
        for f, color in zip(feature_files, colors):
            try:
                df = pd.read_csv(os.path.join(csv_dir, f), header=None, 
                               names=['x', 'y', 'datetime', 'rep'])
                df['feature_id'] = f.replace('.csv', '')
                df['datetime'] = pd.to_datetime(df['datetime'])
                
                ax_paths.plot(df['x'], df['y'], '-', color=color, alpha=0.5, 
                            label=f.replace('.csv', ''))
                
                df['displacement'] = np.sqrt(
                    np.diff(df['x'], append=df['x'].iloc[-1])**2 + 
                    np.diff(df['y'], append=df['y'].iloc[-1])**2
                )
                all_data.append(df)
            except Exception as e:
                print(f"Error processing {f}: {e}")
                continue
        
        ax_paths.set_title('All Feature Paths')
        ax_paths.set_xlabel('X Position (pixels)')
        ax_paths.set_ylabel('Y Position (pixels)')
        ax_paths.grid(True)
        figures['paths'] = fig_paths
        
        if not all_data:
            print("No valid feature data found")
            return figures
            
        # Combine all feature data
        combined_df = pd.concat(all_data, ignore_index=True)
        
        # 2. Activity Heatmap
        fig_heatmap = plt.figure(figsize=(10, 8))
        ax_heatmap = fig_heatmap.add_subplot(111)
        pivot = pd.crosstab(
            combined_df['datetime'].dt.date,
            combined_df['datetime'].dt.hour,
            values=combined_df['displacement'],
            aggfunc='mean'
        ).fillna(0)
        
        sns.heatmap(pivot, cmap='YlOrRd', ax=ax_heatmap)
        ax_heatmap.set_title('Activity Heatmap (Average Displacement)')
        ax_heatmap.set_xlabel('Hour of Day')
        ax_heatmap.set_ylabel('Date')
        figures['heatmap'] = fig_heatmap
        
        # 3. Feature Movement Over Time
        fig_movement = plt.figure(figsize=(10, 8))
        ax_movement = fig_movement.add_subplot(111)
        for name, group in combined_df.groupby('feature_id'):
            ax_movement.plot(group['datetime'], group['displacement'], '-', 
                           alpha=0.5, label=name)
            
        ax_movement.set_title('Feature Movement Over Time')
        ax_movement.set_ylim(0, 100)
        ax_movement.set_xlabel('Time')
        ax_movement.set_ylabel('Displacement (pixels)')
        plt.xticks(rotation=45)
        figures['movement'] = fig_movement
        
        # 4. Daily Statistics
        fig_stats = plt.figure(figsize=(10, 8))
        ax_stats = fig_stats.add_subplot(111)
        daily_stats = combined_df.groupby(combined_df['datetime'].dt.date).agg({
            'feature_id': 'nunique',
            'displacement': 'mean'
        }).reset_index()
        
        ax_stats.bar(range(len(daily_stats)), daily_stats['feature_id'], 
                    alpha=0.5, label='Active Features')
        ax_stats_twin = ax_stats.twinx()
        ax_stats_twin.plot(daily_stats['displacement'], 'r-', label='Avg Movement')
        
        ax_stats.set_title('Daily Activity Summary')
        ax_stats.set_xlabel('Days')
        ax_stats.set_ylabel('Active Features')
        ax_stats_twin.set_ylabel('Average Displacement (pixels)')
        figures['stats'] = fig_stats
        
        for fig in figures.values():
            fig.tight_layout()
            
        # Update path plot with limits
        if plot_limits and 'path' in plot_limits:
            if plot_limits['path']['xmin'] is not None:
                ax_paths.set_xlim(left=plot_limits['path']['xmin'])
            if plot_limits['path']['xmax'] is not None:
                ax_paths.set_xlim(right=plot_limits['path']['xmax'])
            if plot_limits['path']['ymin'] is not None:
                ax_paths.set_ylim(bottom=plot_limits['path']['ymin'])
            if plot_limits['path']['ymax'] is not None:
                ax_paths.set_ylim(top=plot_limits['path']['ymax'])

        if plot_limits and 'movement' in plot_limits:
            if plot_limits['movement']['ymin'] != 0:
                ax_movement.set_ylim(bottom=plot_limits['movement']['ymin'])
            if plot_limits['movement']['ymax'] != 100:
                ax_movement.set_ylim(top=plot_limits['movement']['ymax'])
        
        return figures
        
    except Exception as e:
        print(f"Error creating whole plant visualization: {e}")
        return None

def calculate_path_length(points: np.ndarray):
    """
    Calculate total path length from a sequence of points.

    Args:
        points: numpy array of shape (n,2) containing x,y coordinates

    Returns:
        float: Total path length
    """
    if points is None or len(points) < 2:
        return 0.0

    # Calculate differences between consecutive points
    diffs = np.diff(points, axis=0)
    # Sum up the Euclidean distances (magnitudes of displacement vectors)
    return np.sum(np.sqrt(np.sum(diffs**2, axis=1)))

def compute_vectors(csv_path: str):
    """
    Compute vectors and derived per-step features from CSV file.
    Assumes CSV has columns in order: x, y, datetime, rep

    Args:
        csv_path (str): Path to the CSV file with x,y coordinates and timestamps.

    Returns:
        tuple:
            - vectors (np.ndarray): Array of vectors [x1, y1, dx, dy].
            - coords (np.ndarray): Original (x,y) coordinates.
            - datetimes (list): List of datetime strings.
            - reps (list): List of repetition numbers.
            - magnitudes (np.ndarray): Magnitudes of displacement vectors (speed).
            - angles (np.ndarray): Angles of displacement vectors.
            - acceleration (np.ndarray): Change in magnitudes.
            - jerk (np.ndarray): Change in acceleration.
            - delta_angles (np.ndarray): Change in angles (angular velocity proxy).
            - curvature (np.ndarray): Estimate of path curvature.
    """
    try:
        # Read with header=None and assign names directly to ensure consistency
        df = pd.read_csv(csv_path, header=None, names=['x', 'y', 'datetime', 'rep'])

        coords = df[['x', 'y']].values
        datetimes = df['datetime'].tolist()
        reps = df['rep'].tolist() if 'rep' in df.columns else [np.nan] * len(df)

        if len(coords) < 2:
            # logging.info(f"Not enough data points in {csv_path} to compute vectors (need at least 2).")
            return None, None, None, None, None, None, None, None, None, None

        vectors = np.zeros((len(coords) - 1, 4))  # [x1,y1,dx,dy]
        for i in range(len(coords) - 1):
            vectors[i, 0:2] = coords[i]
            vectors[i, 2:4] = coords[i+1] - coords[i]

        # Derived per-step features
        magnitudes = np.linalg.norm(vectors[:, 2:4], axis=1) # Speed
        angles = np.rad2deg(np.arctan2(vectors[:, 3], vectors[:, 2])) # Angle in degrees
        angles = (angles + 360) % 360 # Normalize angles to [0, 360)

        acceleration = np.diff(magnitudes, prepend=magnitudes[0]) # Change in speed
        jerk = np.diff(acceleration, prepend=acceleration[0]) # Change in acceleration

        # Delta angles (angular velocity) - handling wrap-around for correct change
        raw_delta_angles = np.diff(angles, prepend=angles[0])
        # Adjust for angles crossing 0/360 boundary
        delta_angles = np.where(raw_delta_angles > 180, raw_delta_angles - 360, raw_delta_angles)
        delta_angles = np.where(delta_angles < -180, delta_angles + 360, delta_angles)

        # Curvature: ratio of angle change to magnitude (large turn, small step = high curvature)
        curvature = np.zeros_like(magnitudes)
        # Avoid division by zero: only calculate for steps with non-zero magnitude
        non_zero_mag_idx = magnitudes > 0.001 # Small threshold
        curvature[non_zero_mag_idx] = np.abs(delta_angles[non_zero_mag_idx]) / magnitudes[non_zero_mag_idx]

        return vectors, coords, datetimes, reps, magnitudes, angles, \
               acceleration, jerk, delta_angles, curvature

    except Exception as e:
        print(f"Error reading or processing {csv_path}: {e}")
        return None, None, None, None, None, None, None, None, None, None

def compute_resultant_vector(vectors: np.ndarray):
    """Compute resultant vector from a set of vectors."""
    if vectors is None or len(vectors) == 0:
        return np.array([np.nan, np.nan])
    resultant_x = np.sum(vectors[:, 2])
    resultant_y = np.sum(vectors[:, 3])
    return np.array([resultant_x, resultant_y])

def filter_outliers(data, method='iqr', threshold=1.5):
    """
    Filters outliers from a 1D numpy array or pandas Series.

    Args:
        data (np.ndarray or pd.Series): The input data.
        method (str): 'iqr' for Interquartile Range or 'std' for standard deviation.
        threshold (float): Multiplier for IQR or std deviation.

    Returns:
        np.ndarray: Data with outliers removed.
    """
    if isinstance(data, pd.Series):
        data = data.values

    if not isinstance(data, np.ndarray) or data.ndim != 1:
        raise ValueError("Input 'data' must be a 1D numpy array or pandas Series.")

    if len(data) == 0:
        return np.array([])

    data = data[~np.isnan(data)] # Remove NaNs first

    if len(data) == 0:
        return np.array([])

    if method == 'iqr':
        Q1 = np.percentile(data, 25)
        Q3 = np.percentile(data, 75)
        IQR = Q3 - Q1
        lower_bound = Q1 - threshold * IQR
        upper_bound = Q3 + threshold * IQR
    elif method == 'std':
        data_mean = np.mean(data)
        data_std = np.std(data)
        lower_bound = data_mean - threshold * data_std
        upper_bound = data_mean + threshold * data_std
    else:
        raise ValueError("Method must be 'iqr' or 'std'.")

    filtered_data = data[(data >= lower_bound) & (data <= upper_bound)]
    return filtered_data

def compute_advanced_metrics(vectors: np.ndarray, datetimes: list, magnitudes: np.ndarray, angles: np.ndarray, delta_angles: np.ndarray):
    """
    Computes advanced motion metrics: smoothness, direction consistency, and complexity.

    Args:
        vectors (np.ndarray): Array of vectors [x1, y1, dx, dy].
        datetimes (list): List of datetime strings corresponding to the points.
        magnitudes (np.ndarray): Pre-computed magnitudes.
        angles (np.ndarray): Pre-computed angles.
        delta_angles (np.ndarray): Pre-computed delta_angles.

    Returns:
        tuple: (smoothness, direction_consistency, complexity, tortuosity_index)
    """
    if vectors is None or len(vectors) < 2:
        return np.nan, np.nan, np.nan, np.nan

    # Smoothness: Standard deviation of (filtered) angular changes (lower std dev means smoother)
    filtered_delta_angles = filter_outliers(np.abs(delta_angles), method='iqr', threshold=2.0)
    smoothness = np.std(filtered_delta_angles) if len(filtered_delta_angles) > 1 else np.nan

    # Direction Consistency: Ratio of resultant vector magnitude to sum of individual magnitudes.
    resultant_vec = compute_resultant_vector(vectors)
    total_path_mag = np.sum(magnitudes) # This is the path length
    if total_path_mag > 0:
        direction_consistency = np.linalg.norm(resultant_vec) / total_path_mag
    else:
        direction_consistency = np.nan

    # Complexity: Number of "significant turns" (e.g., angle change > 10 degrees)
    complexity_turns = np.sum(np.abs(delta_angles) > 10) # Count instances where angle changes by more than 10 degrees

    # Tortuosity Index: Ratio of path length to net displacement
    if len(vectors) > 0:
        start_point = vectors[0, :2] # x1, y1 of the first vector
        end_point = vectors[-1, :2] + vectors[-1, 2:] # (x1+dx, y1+dy) of the last vector
        net_displacement = np.linalg.norm(end_point - start_point)
    else:
        net_displacement = 0.0

    if net_displacement > 0:
        tortuosity_index = total_path_mag / net_displacement
    else:
        tortuosity_index = np.nan # If no net displacement, tortuosity is undefined or infinite

    return smoothness, direction_consistency, complexity_turns, tortuosity_index

def create_motion_visualizations(points: np.ndarray, datetimes: list):
    """
    Creates various motion visualizations and returns them as BytesIO objects.

    Args:
        points (np.ndarray): Array of x,y coordinates.
        datetimes (list): List of datetime strings corresponding to the points.

    Returns:
        dict: A dictionary of BytesIO objects for each plot.
    """
    plots = {}

    # Check for insufficient data for any plot
    if points is None or len(points) < 2:
        logging.warning("Insufficient points data for visualization. Returning placeholder plots.")
        # Trajectory plot placeholder
        fig, ax = plt.subplots(figsize=(8, 6))
        ax.text(0.5, 0.5, 'No valid data for trajectory plot', horizontalalignment='center', verticalalignment='center', transform=ax.transAxes)
        fig.tight_layout()
        buf = io.BytesIO()
        fig.savefig(buf, format='png')
        plt.close(fig)
        buf.seek(0)
        plots['trajectory_plot'] = buf

        # Velocity plot placeholder
        fig, ax = plt.subplots(figsize=(8, 6))
        ax.text(0.5, 0.5, 'No valid data for velocity plot', horizontalalignment='center', verticalalignment='center', transform=ax.transAxes)
        fig.tight_layout()
        buf = io.BytesIO()
        fig.savefig(buf, format='png')
        plt.close(fig)
        buf.seek(0)
        plots['velocity_time_series'] = buf

        # Animation first frame placeholder
        plots['animation_first_frame'] = None # Or a placeholder image if preferred
        return plots

    # --- Trajectory Plot ---
    fig_traj, ax_traj = plt.subplots(figsize=(8, 6))
    ax_traj.plot(points[:, 0], points[:, 1], marker='o', linestyle='-', markersize=2, alpha=0.7)
    ax_traj.set_title("Movement Trajectory")
    ax_traj.set_xlabel("X Coordinate")
    ax_traj.set_ylabel("Y Coordinate")
    ax_traj.grid(True)
    ax_traj.set_aspect('equal', adjustable='box') # Ensure correct aspect ratio
    fig_traj.tight_layout()
    buf_traj = io.BytesIO()
    fig_traj.savefig(buf_traj, format='png')
    plt.close(fig_traj)
    buf_traj.seek(0)
    plots['trajectory_plot'] = buf_traj

    # --- Velocity Time Series Plot (using magnitudes) ---
    magnitudes_for_plot = np.linalg.norm(np.diff(points, axis=0), axis=1)

    # Convert datetimes to a pandas Series *explicitly* to ensure .iloc is available
    datetime_series_pd = pd.Series(pd.to_datetime(datetimes, errors='coerce'))

    # Use the timestamps corresponding to the *end* of the interval for plotting velocity
    # This means we use datetimes from the second point onwards.
    if len(datetime_series_pd) > 1:
        time_points = datetime_series_pd.iloc[1:].values
    else:
        time_points = np.array([]) # No time points if less than 2 datetimes

    fig_vel, ax_vel = plt.subplots(figsize=(10, 6))
    if len(time_points) == len(magnitudes_for_plot) and len(time_points) > 0:
        ax_vel.plot(time_points, magnitudes_for_plot, linestyle='-', marker='.', markersize=4)
        ax_vel.set_title("Velocity (Magnitude) Over Time")
        ax_vel.set_xlabel("Time")
        ax_vel.set_ylabel("Velocity (pixels/time unit)")
        ax_vel.grid(True)
        fig_vel.autofmt_xdate()
    else:
        logging.warning("Mismatch in data lengths or insufficient data for velocity plot. Plotting placeholder.")
        ax_vel.text(0.5, 0.5, 'Mismatch in data lengths or insufficient data for velocity plot', horizontalalignment='center', verticalalignment='center', transform=ax_vel.transAxes)

    fig_vel.tight_layout()
    buf_vel = io.BytesIO()
    fig_vel.savefig(buf_vel, format='png')
    plt.close(fig_vel)
    buf_vel.seek(0)
    plots['velocity_time_series'] = buf_vel

    # --- Motion Animation (First Frame) ---
    fig_anim, ax_anim = plt.subplots(figsize=(8, 8))
    ax_anim.set_title("Movement Animation (First Frame)")
    ax_anim.set_xlabel("X")
    ax_anim.set_ylabel("Y")

    if len(points) > 0:
        x_min, x_max = np.min(points[:,0]), np.max(points[:,0])
        y_min, y_max = np.min(points[:,1]), np.max(points[:,1])
        # Add some padding
        x_range = x_max - x_min
        y_range = y_max - y_min
        ax_anim.set_xlim(x_min - x_range*0.1, x_max + x_range*0.1)
        ax_anim.set_ylim(y_min - y_range*0.1, y_max + y_range*0.1)
        ax_anim.set_aspect('equal', adjustable='box')

        ax_anim.plot(points[0,0], points[0,1], 'ro', markersize=8, label='Start Point') # Start point
        if len(points) > 1:
            ax_anim.plot(points[:,0], points[:,1], 'b-', alpha=0.5, label='Trajectory') # Entire trajectory
            ax_anim.plot(points[-1,0], points[-1,1], 'go', markersize=8, label='End Point') # End point
        ax_anim.legend()
    else:
        logging.warning("No points data for animation frame. Plotting placeholder.")
        ax_anim.text(0.5, 0.5, 'No points data for animation frame', horizontalalignment='center', verticalalignment='center', transform=ax_anim.transAxes)

    buf_anim = io.BytesIO()
    fig_anim.tight_layout()
    fig_anim.savefig(buf_anim, format='png')
    plt.close(fig_anim)
    buf_anim.seek(0)
    plots['animation_first_frame'] = buf_anim

    return plots

def analyze_motion_patterns(vectors: np.ndarray, datetimes: list, magnitudes: np.ndarray):
    """
    Analyzes temporal motion patterns (hourly, daily, frequency).

    Args:
        vectors (np.ndarray): Array of vectors [x1, y1, dx, dy].
        datetimes (list): List of datetime strings corresponding to the points.
        magnitudes (np.ndarray): Pre-computed magnitudes.

    Returns:
        tuple: (df_hourly_patterns, df_daily_patterns, df_frequency_analysis)
    """
    if vectors is None or len(vectors) == 0:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    # datetimes[:-1] because magnitudes are for the steps between points
    datetime_series_aligned = pd.to_datetime(datetimes[:-1], errors='coerce')

    df_temp = pd.DataFrame({
        'magnitude': magnitudes,
        'datetime': datetime_series_aligned
    }).dropna(subset=['datetime'])

    if df_temp.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    # Hourly patterns: Mean magnitude per hour of day
    df_temp['hour'] = df_temp['datetime'].dt.hour
    hourly_patterns = df_temp.groupby('hour')['magnitude'].mean().reset_index()
    hourly_patterns.rename(columns={'magnitude': 'mean_magnitude'}, inplace=True)

    # Daily patterns: Mean magnitude per day of week
    df_temp['day_of_week'] = df_temp['datetime'].dt.day_name()
    day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    # Use categorical type for correct sorting and handling of missing days
    df_temp['day_of_week'] = pd.Categorical(df_temp['day_of_week'], categories=day_order, ordered=True)
    daily_patterns = df_temp.groupby('day_of_week')['magnitude'].mean().reindex(day_order).reset_index()
    daily_patterns.rename(columns={'magnitude': 'mean_magnitude'}, inplace=True)
    daily_patterns.dropna(inplace=True) # Drop days that had no data

    # Frequency analysis: Using FFT on magnitudes
    frequency_analysis_df = pd.DataFrame()
    if len(df_temp) > 1:
        try:
            magnitudes_for_fft = df_temp['magnitude'].values
            N = len(magnitudes_for_fft)

            # Calculate the average time step
            time_diffs_seconds = np.diff(df_temp['datetime'].apply(lambda x: x.timestamp()))
            T = np.mean(time_diffs_seconds) if len(time_diffs_seconds) > 0 else 1.0 # Average time step in seconds

            if N > 0 and T > 0:
                yf = np.fft.fft(magnitudes_for_fft)
                xf = np.fft.fftfreq(N, T)[:N//2] # Frequencies in Hz
                frequency_analysis_df = pd.DataFrame({
                    'frequency_Hz': xf,
                    'amplitude': 2.0/N * np.abs(yf[0:N//2]) # Amplitude spectrum
                })
        except Exception as e:
            print(f"Warning: Frequency analysis failed for magnitudes: {e}")

    return hourly_patterns, daily_patterns, frequency_analysis_df


# --- NEW: Function to calculate and export aggregated plant motion ---

def calculate_and_export_aggregated_motion(
    motion_features_dir: str,
    overall_datetimes_agg: list[str],
    output_filename: str = "aggregated_plant_motion.csv"
):
    """
    Calculates aggregated motion metrics for the entire plant over time.
    This aggregates data from all individual feature CSVs at each time point
    specified by overall_datetimes_agg.

    Args:
        motion_features_dir (str): Directory containing the feature_*.csv files.
        overall_datetimes_agg (list[str]): Sorted list of all datetime strings
                                            from the image sequence.
        output_filename (str): Name of the CSV file to save aggregated data.

    Returns:
        str: Path to the generated aggregated motion CSV file, or None if failed.
    """
    print(f"AGGREGATION: Starting aggregation for {motion_features_dir}")
    if not os.path.isdir(motion_features_dir):
        print(f"AGGREGATION ERROR: Motion features directory not found: {motion_features_dir}")
        return None

    if not overall_datetimes_agg:
        print("AGGREGATION WARNING: No datetimes provided for aggregation. Skipping.")
        return None

    # Convert overall_datetimes_agg to a sorted list of pandas Timestamps
    # This serves as our master timeline for the plant's activity
    master_timeline_dt = pd.to_datetime(natsorted(list(set(overall_datetimes_agg))), errors='coerce')
    master_timeline_dt = master_timeline_dt.dropna().sort_values().unique() # Remove NaT and ensure uniqueness/sort

    if len(master_timeline_dt) < 2:
        print("AGGREGATION WARNING: Insufficient valid datetimes in timeline for aggregation.")
        return None

    # Initialize a DataFrame to store aggregated data
    # Use the master timeline as the index
    aggregated_data = pd.DataFrame(index=master_timeline_dt)
    aggregated_data.index.name = 'datetime'

    # Columns for aggregated metrics
    aggregated_data['active_features_count'] = 0
    aggregated_data['mean_magnitude_active_features'] = np.nan
    aggregated_data['mean_acceleration_active_features'] = np.nan
    aggregated_data['mean_jerk_active_features'] = np.nan
    aggregated_data['mean_delta_angle_active_features'] = np.nan


    feature_csv_files = [
        os.path.join(motion_features_dir, f)
        for f in os.listdir(motion_features_dir)
        if f.startswith('feature_') and f.lower().endswith('.csv')
    ]
    feature_csv_files = natsorted(feature_csv_files)

    all_feature_dfs = {} # Store individual feature dataframes for efficient lookup

    print(f"AGGREGATION: Found {len(feature_csv_files)} individual feature CSVs.")

    for f_path in feature_csv_files:
        feature_id_match = re.search(r'feature_(\d+)\.csv', os.path.basename(f_path))
        if not feature_id_match:
            continue
        feature_id = int(feature_id_match.group(1))

        try:
            # Read with header=None and assign names
            df_feature = pd.read_csv(f_path, header=None, names=['x', 'y', 'datetime', 'rep'])
            df_feature['datetime'] = pd.to_datetime(df_feature['datetime'], errors='coerce')
            df_feature = df_feature.dropna(subset=['datetime']).set_index('datetime').sort_index()

            if not df_feature.empty:
                all_feature_dfs[feature_id] = df_feature
            # else:
            #     print(f"WARNING: Feature {feature_id} CSV is empty or has no valid datetimes.")
        except Exception as e:
            print(f"AGGREGATION WARNING: Could not read or process feature CSV {f_path}: {e}")

    # Iterate through the master timeline to aggregate metrics
    # We aggregate *between* time points, similar to how magnitudes are calculated.
    # So, for each interval (t_i, t_{i+1}), we look at features active at t_i and their movement to t_{i+1}
    
    # Let's adjust aggregated_data to have a range of dates instead of single points
    # This makes the most sense if we're measuring motion *between* frames.
    # The first 'motion' event happens between the first and second frames.

    aggregated_list = []

    for i in range(len(master_timeline_dt) - 1):
        current_time = master_timeline_dt[i]
        next_time = master_timeline_dt[i+1]
        
        active_features = []
        magnitudes_in_interval = []
        accelerations_in_interval = []
        jerks_in_interval = []
        delta_angles_in_interval = []

        # Iterate through all loaded features to find their state in this interval
        for feature_id, df_feature in all_feature_dfs.items():
            # Find the point for the current time
            current_pos = df_feature.loc[df_feature.index == current_time]
            next_pos = df_feature.loc[df_feature.index == next_time]

            if not current_pos.empty and not next_pos.empty:
                # Feature was present at current_time and next_time, so it moved
                x1, y1 = current_pos.iloc[0][['x', 'y']].values
                x2, y2 = next_pos.iloc[0][['x', 'y']].values

                dx = x2 - x1
                dy = y2 - y1
                magnitude = np.linalg.norm([dx, dy])

                # To get acceleration, jerk, delta_angle, we need more context.
                # For simplicity in this aggregated view, we'll just use magnitude directly.
                # If we need true per-feature acceleration/jerk, `compute_vectors` is better
                # to call on the full feature path and then extract relevant segments.
                # For plant-level aggregation, usually 'activity' (magnitudes) is key.

                # Let's re-use compute_vectors to get these if needed, but it's computationally
                # intensive to do for every interval for every feature.
                # A simpler approach: if a feature moved, it's 'active'.
                
                # For plant-level, let's just count active features and mean magnitude.
                # Other detailed metrics are best from the individual feature CSV analysis.

                if magnitude > 0: # Consider a feature active if it moved
                    active_features.append(feature_id)
                    magnitudes_in_interval.append(magnitude)
                    
                    # For a simple aggregated acceleration/jerk, we'd need speed history.
                    # This is complex to do across features for arbitrary intervals.
                    # We'll calculate simple mean/max of active magnitudes for now.

        num_active = len(active_features)
        mean_mag = np.mean(magnitudes_in_interval) if magnitudes_in_interval else 0.0 # Use 0.0 for intervals with no motion
        
        aggregated_list.append({
            'datetime_start': current_time.strftime("%m/%d/%Y %H:%M:%S"),
            'datetime_end': next_time.strftime("%m/%d/%Y %H:%M:%S"),
            'active_features_count': num_active,
            'total_magnitude_sum': np.sum(magnitudes_in_interval) if magnitudes_in_interval else 0.0,
            'mean_magnitude_active_features': mean_mag
            # Add other aggregated metrics here if derived from magnitudes_in_interval
        })

    df_aggregated = pd.DataFrame(aggregated_list)
    
    # Save the aggregated data to a CSV file
    output_csv_path = os.path.join(motion_features_dir, output_filename)
    try:
        df_aggregated.to_csv(output_csv_path, index=False)
        print(f"AGGREGATION: Aggregated plant motion data saved to: {output_csv_path}")
        return output_csv_path
    except Exception as e:
        print(f"AGGREGATION ERROR: Could not save aggregated motion CSV to {output_csv_path}: {e}")
        return None

# --- Processing and Export Functions (Modified) ---

def batch_process(csv_files: list[str]):
    """Process multiple CSV files and collect all relevant motion metrics."""
    # Lists to store aggregated results across all files
    all_file_summaries = []

    # Data for individual file details (for per-file sheets in Excel)
    per_file_detailed_data = []

    for csv_file in natsorted(csv_files):
        # Initial call to compute_vectors to get all per-step features
        vectors, points, datetimes, reps, magnitudes, angles, acceleration, jerk, delta_angles, curvature = \
            compute_vectors(csv_file)

        file_name = os.path.basename(csv_file)

        if vectors is None or len(vectors) == 0:
            print(f"Skipping {file_name}: No valid data.")
            # Append NaN or empty dict for this file to maintain structure
            all_file_summaries.append({
                'file': file_name,
                'mean_speed': np.nan, 'max_speed': np.nan,
                'mean_acceleration': np.nan, 'max_acceleration': np.nan,
                'mean_jerk': np.nan, 'max_jerk': np.nan,
                'mean_abs_angular_velocity': np.nan, 'num_significant_turns': np.nan,
                'path_length': np.nan, 'convex_hull_area': np.nan,
                'smoothness': np.nan, 'direction_consistency': np.nan,
                'complexity_turns': np.nan, 'tortuosity_index': np.nan,
                'resultant_x': np.nan, 'resultant_y': np.nan,
                'first_datetime': pd.NaT, 'last_datetime': pd.NaT, 'duration_seconds': np.nan
            })
            per_file_detailed_data.append({
                'file': file_name,
                'vectors': None, 'points': None, 'datetimes': None, 'reps': None,
                'magnitudes': None, 'angles': None, 'euclidean_distances': None,
                'acceleration': None, 'jerk': None, 'delta_angles': None, 'curvature': None,
                'daily_path_lengths': [] # Ensure this is an empty list
            })
            continue

        # Convert datetimes to pandas datetime objects for easier manipulation
        datetime_series = pd.to_datetime(datetimes, errors='coerce')
        first_datetime = datetime_series.min()
        last_datetime = datetime_series.max()
        duration_seconds = (last_datetime - first_datetime).total_seconds() if pd.notna(first_datetime) and pd.notna(last_datetime) else np.nan


        # Ensure arrays are not empty before calculations
        if len(magnitudes) == 0:
            print(f"Skipping calculations for {file_name}: Empty magnitudes array.")
            # Append NaN or empty dict for this file to maintain structure
            all_file_summaries.append({
                'file': file_name,
                'mean_speed': np.nan, 'max_speed': np.nan,
                'mean_acceleration': np.nan, 'max_acceleration': np.nan,
                'mean_jerk': np.nan, 'max_jerk': np.nan,
                'mean_abs_angular_velocity': np.nan, 'num_significant_turns': np.nan,
                'path_length': np.nan, 'convex_hull_area': np.nan,
                'smoothness': np.nan, 'direction_consistency': np.nan,
                'complexity_turns': np.nan, 'tortuosity_index': np.nan,
                'resultant_x': np.nan, 'resultant_y': np.nan,
                'first_datetime': first_datetime, 'last_datetime': last_datetime, 'duration_seconds': duration_seconds
            })
            per_file_detailed_data.append({
                'file': file_name,
                'vectors': None, 'points': None, 'datetimes': None, 'reps': None,
                'magnitudes': None, 'angles': None, 'euclidean_distances': None,
                'acceleration': None, 'jerk': None, 'delta_angles': None, 'curvature': None,
                'daily_path_lengths': []
            })
            continue

        # --- Aggregate Metrics for this file ---
        mean_speed = np.mean(magnitudes)
        max_speed = np.max(magnitudes)

        mean_acceleration = np.mean(np.abs(acceleration))
        max_acceleration = np.max(np.abs(acceleration))

        mean_jerk = np.mean(np.abs(jerk))
        max_jerk = np.max(np.abs(jerk))

        mean_abs_delta_angle = np.mean(np.abs(delta_angles))
        num_significant_turns = np.sum(np.abs(delta_angles) > 10)

        path_length = calculate_path_length(points)

        # Convex Hull Area
        convex_hull_area = np.nan
        if len(points) >= 3:
            try:
                hull = ConvexHull(points)
                convex_hull_area = hull.area
            except Exception as e:
                # This can happen if all points are collinear
                # print(f"Warning: Could not compute convex hull for {file_name}: {e}")
                pass # Remain NaN

        smoothness, direction_consistency, complexity_turns, tortuosity_index = \
            compute_advanced_metrics(vectors, datetimes, magnitudes, angles, delta_angles)

        resultant_vec = compute_resultant_vector(vectors)

        # Daily Path Length Calculation for the current file
        daily_path_lengths_for_file = []
        try:
            # Only consider valid datetimes for grouping
            valid_df = pd.DataFrame({
                'x': points[:, 0],
                'y': points[:, 1],
                'datetime': datetime_series
            }).dropna(subset=['datetime']) # Drop rows where datetime couldn't be parsed

            if not valid_df.empty:
                valid_df['date'] = valid_df['datetime'].dt.date
                grouped = valid_df.groupby('date')
                for date, group in grouped:
                    coords = group[['x', 'y']].to_numpy()
                    if len(coords) > 1:
                        daily_len = calculate_path_length(coords)
                        daily_path_lengths_for_file.append({'file': file_name, 'date': date.strftime('%Y-%m-%d'), 'daily_path_length': daily_len})
        except Exception as e:
            print(f"Skipping daily path length for {file_name} due to error: {e}")


        # Store summary for batch output
        all_file_summaries.append({
            'file': file_name,
            'mean_speed': mean_speed, 'max_speed': max_speed,
            'mean_acceleration': mean_acceleration, 'max_acceleration': max_acceleration,
            'mean_jerk': mean_jerk, 'max_jerk': max_jerk,
            'mean_abs_angular_velocity': mean_abs_delta_angle,
            'num_significant_turns': num_significant_turns,
            'path_length': path_length,
            'convex_hull_area': convex_hull_area,
            'smoothness': smoothness,
            'direction_consistency': direction_consistency,
            'complexity_turns': complexity_turns,
            'tortuosity_index': tortuosity_index,
            'resultant_x': resultant_vec[0],
            'resultant_y': resultant_vec[1],
            'first_datetime': first_datetime,
            'last_datetime': last_datetime,
            'duration_seconds': duration_seconds
        })

        # Store all detailed per-step data for individual file sheets
        per_file_detailed_data.append({
            'file': file_name,
            'vectors': vectors,
            'points': points,
            'datetimes': datetimes, # Original datetime strings
            'reps': reps,
            'magnitudes': magnitudes,
            'angles': angles,
            'euclidean_distances': magnitudes, # Magnitudes are the Euclidean distances of each step
            'acceleration': acceleration,
            'jerk': jerk,
            'delta_angles': delta_angles,
            'curvature': curvature,
            'daily_path_lengths': daily_path_lengths_for_file # Store this for each file
        })

    # Consolidate all daily path lengths into a single DataFrame for the "AllDailyPathLengths" sheet
    df_all_daily_paths = pd.DataFrame([item for sublist in [d['daily_path_lengths'] for d in per_file_detailed_data if d is not None] for item in sublist])

    return pd.DataFrame(all_file_summaries), df_all_daily_paths, per_file_detailed_data

def single_process(csv_path: str):
    """Process a single CSV file and return all relevant metrics."""
    if not csv_path:
        print("No file selected.")
        return None, pd.DataFrame(), []

    vectors, points, datetimes, reps, magnitudes, angles, acceleration, jerk, delta_angles, curvature = \
        compute_vectors(csv_path)

    if vectors is None or len(vectors) == 0:
        print(f"No valid data in {csv_path}")
        return None, pd.DataFrame(), []

    file_name = os.path.basename(csv_path)

    # Convert datetimes to pandas datetime objects for easier manipulation
    datetime_series = pd.to_datetime(datetimes, errors='coerce')
    first_datetime = datetime_series.min()
    last_datetime = datetime_series.max()
    duration_seconds = (last_datetime - first_datetime).total_seconds() if pd.notna(first_datetime) and pd.notna(last_datetime) else np.nan

    # --- Aggregate Metrics for this file ---
    mean_speed = np.mean(magnitudes)
    max_speed = np.max(magnitudes)

    mean_acceleration = np.mean(np.abs(acceleration))
    max_acceleration = np.max(np.abs(acceleration))

    mean_jerk = np.mean(np.abs(jerk))
    max_jerk = np.max(np.abs(jerk))

    mean_abs_delta_angle = np.mean(np.abs(delta_angles))
    num_significant_turns = np.sum(np.abs(delta_angles) > 10)

    path_length = calculate_path_length(points)

    # Convex Hull Area
    convex_hull_area = np.nan
    if len(points) >= 3:
        try:
            hull = ConvexHull(points)
            convex_hull_area = hull.area
        except Exception as e:
            pass # Remain NaN

    smoothness, direction_consistency, complexity_turns, tortuosity_index = \
        compute_advanced_metrics(vectors, datetimes, magnitudes, angles, delta_angles)

    resultant_vec = compute_resultant_vector(vectors)

    # Daily Path Length Calculation for the current file
    daily_path_lengths_for_file = []
    try:
        valid_df = pd.DataFrame({
            'x': points[:, 0],
            'y': points[:, 1],
            'datetime': datetime_series
        }).dropna(subset=['datetime'])

        if not valid_df.empty:
            valid_df['date'] = valid_df['datetime'].dt.date
            grouped = valid_df.groupby('date')
            for date, group in grouped:
                coords = group[['x', 'y']].to_numpy()
                if len(coords) > 1:
                    daily_len = calculate_path_length(coords)
                    daily_path_lengths_for_file.append({'file': file_name, 'date': date.strftime('%Y-%m-%d'), 'daily_path_length': daily_len})
    except Exception as e:
        print(f"Skipping daily path length for {file_name} due to error: {e}")


    # Prepare a dictionary for summary (similar to what batch_process returns for one file)
    summary_data = {
        'file': file_name,
        'mean_speed': mean_speed, 'max_speed': max_speed,
        'mean_acceleration': mean_acceleration, 'max_acceleration': max_acceleration,
        'mean_jerk': mean_jerk, 'max_jerk': max_jerk,
        'mean_abs_angular_velocity': mean_abs_delta_angle,
        'num_significant_turns': num_significant_turns,
        'path_length': path_length,
        'convex_hull_area': convex_hull_area,
        'smoothness': smoothness,
        'direction_consistency': direction_consistency,
        'complexity_turns': complexity_turns,
        'tortuosity_index': tortuosity_index,
        'resultant_x': resultant_vec[0],
        'resultant_y': resultant_vec[1],
        'first_datetime': first_datetime,
        'last_datetime': last_datetime,
        'duration_seconds': duration_seconds
    }

    # Prepare detailed data (similar structure to per_file_detailed_data from batch_process)
    detailed_data = {
        'file': file_name,
        'vectors': vectors,
        'points': points,
        'datetimes': datetimes,
        'reps': reps,
        'magnitudes': magnitudes,
        'angles': angles,
        'euclidean_distances': magnitudes,
        'acceleration': acceleration,
        'jerk': jerk,
        'delta_angles': delta_angles,
        'curvature': curvature,
        'daily_path_lengths': daily_path_lengths_for_file
    }

    return pd.DataFrame([summary_data]), pd.DataFrame(daily_path_lengths_for_file), [detailed_data]

def export_to_excel(
    batch_mode: bool = False,
    csv_dir: str = None,
    csv_path: str = None,
    output_filename: str = "motion_analysis_report.xlsx",
    aggregated_motion_csv_path: str = None, # NEW PARAMETER
    plot_limits: dict = None
):
    """
    Exports motion analysis results to an Excel file with detailed sheets and plots.

    Parameters:
    batch_mode (bool): If True, processes all CSVs in csv_dir.
    csv_dir (str): Directory containing CSV files (for batch_mode).
    csv_path (str): Path to a single CSV file (for non-batch mode).
    output_filename (str): Name of the Excel file to generate.
    aggregated_motion_csv_path (str, optional): Path to the aggregated plant motion CSV.
                                                 If provided, its data will be added to a new sheet.
    """
    print("EXPORT: Entering export_to_excel function.")
    print(f"EXPORT: Initial batch_mode: {batch_mode}, csv_dir: {csv_dir}, csv_path: {csv_path}")

    input_csv_files = []
    excel_output_path = None

    if batch_mode:
        if not csv_dir or not os.path.isdir(csv_dir):
            print("EXPORT ERROR: CSV directory not provided or invalid for batch mode.")
            return
        print(f"EXPORT: Batch mode activated for directory: {csv_dir}")
        input_csv_files = [os.path.join(csv_dir, f) for f in os.listdir(csv_dir) if f.lower().startswith('feature_') and f.lower().endswith('.csv')]
        input_csv_files = natsorted(input_csv_files) # Ensure sorted order
        excel_output_dir = csv_dir
        excel_output_path = os.path.join(excel_output_dir, output_filename)
    elif csv_path:
        if not os.path.isfile(csv_path):
            print(f"EXPORT ERROR: CSV file not found at: {csv_path}")
            return
        print(f"EXPORT: Single file mode for: {csv_path}")
        input_csv_files.append(csv_path)
        excel_output_dir = os.path.dirname(csv_path)
        excel_output_path = os.path.join(excel_output_dir, output_filename)
    else:
        print("EXPORT ERROR: No valid input provided (csv_dir or csv_path).")
        return
    

    if not input_csv_files:
        print("EXPORT WARNING: No individual feature CSV files found to process for detailed sheets. Proceeding with aggregated data if available.")
        # We might still proceed if only aggregated_motion_csv_path is provided.
        if not aggregated_motion_csv_path or not os.path.isfile(aggregated_motion_csv_path):
            print("EXPORT WARNING: No feature CSVs and no aggregated CSV found. Exiting export.")
            return

    # Process data based on mode for individual feature CSVs
    df_summary, df_all_daily_paths, all_detailed_data = pd.DataFrame(), pd.DataFrame(), []
    if input_csv_files: # Only call process if there are feature CSVs
        if batch_mode:
            df_summary, df_all_daily_paths, all_detailed_data = batch_process(input_csv_files)
        else:
            df_summary, df_all_daily_paths, all_detailed_data = single_process(input_csv_files[0])
            if df_summary is None: # single_process returns None for failed cases
                print("EXPORT ERROR: Single feature file processing failed. No detailed sheets will be generated.")
                df_summary = pd.DataFrame() # Ensure it's an empty DataFrame

    if df_summary.empty and not aggregated_motion_csv_path:
        print("EXPORT WARNING: No data processed to generate a report. Exiting export.")
        return

    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    if batch_mode:
        figures = create_whole_plant_visualization(csv_dir, plot_limits=plot_limits)
        ws_plant = wb.create_sheet("Whole_Plant_Analysis")
        current_row = 1
        
        if figures: # Check if figures were successfully created
            for plot_name, fig in figures.items():
                if fig: # Ensure the figure object is not None
                    # Save each visualization
                    img_data = io.BytesIO()
                    fig.savefig(img_data, format='png', dpi=300, bbox_inches='tight')
                    plt.close(fig)
                    img_data.seek(0)
                    
                    # Add title for each plot
                    ws_plant.cell(row=current_row, column=1, value=f"{plot_name.title()} Plot")
                    current_row += 1
                    
                    # Add image
                    img = openpyxl.drawing.image.Image(img_data)
                    ws_plant.add_image(img, f'A{current_row}')
                    current_row += 30  # Adjust spacing between plots
        else:
            print("EXPORT WARNING: No whole plant visualization figures generated for batch mode.")


        # --- Handle Circadian Analysis Plots (Batch Mode) ---
        hourly_activity, cwtmatr, widths = analyze_circadian_rhythm(csv_dir)
        
        # Unpack the two figures returned by plot_circadian_analysis
        circ_main_fig = plot_circadian_analysis(hourly_activity, cwtmatr, widths, plot_limits=plot_limits)
        
        ws_circadian = wb.create_sheet("Circadian_Analysis")
        
        circadian_current_row = 1
        
        if circ_main_fig:
            ws_circadian.cell(row=circadian_current_row, column=1, value="Wavelet Scalogram and Activity (Main)")
            circadian_current_row += 1
            img_data_main = io.BytesIO()
            circ_main_fig.savefig(img_data_main, format='png', dpi=300, bbox_inches='tight')
            plt.close(circ_main_fig)
            img_data_main.seek(0)
            img_main = openpyxl.drawing.image.Image(img_data_main)
            ws_circadian.add_image(img_main, f'A{circadian_current_row}')
            circadian_current_row += 30 # Adjust spacing


        if not circ_main_fig:
            print("EXPORT WARNING: No circadian analysis figures generated for batch mode.")


            # Add displacement data sheet
        displacement_sheet = wb.create_sheet(title="Displacement Data")
        displacement_sheet.append(["Datetime", "Displacement"])
        for index, row in hourly_activity.iterrows():
            displacement_sheet.append([index, row['displacement']])
            
        displacement_sheet.column_dimensions['A'].width = 23
        


    else: # Single file mode
        figures = create_whole_plant_visualization(os.path.dirname(csv_path), plot_limits=plot_limits)
        ws_plant = wb.create_sheet("Whole_Plant_Analysis")
        current_row = 1
        
        if figures: # Check if figures were successfully created
            for plot_name, fig in figures.items():
                if fig: # Ensure the figure object is not None
                    # Save each visualization
                    img_data = io.BytesIO()
                    fig.savefig(img_data, format='png', dpi=300, bbox_inches='tight')
                    plt.close(fig)
                    img_data.seek(0)
                    
                    # Add title for each plot
                    ws_plant.cell(row=current_row, column=1, value=f"{plot_name.title()} Plot")
                    current_row += 1
                    
                    # Add image
                    img = openpyxl.drawing.image.Image(img_data)
                    ws_plant.add_image(img, f'A{current_row}')
                    current_row += 30  # Adjust spacing between plots
        else:
            print("EXPORT WARNING: No whole plant visualization figures generated for single mode.")

        # --- Handle Circadian Analysis Plots (Single File Mode) ---
        hourly_activity, cwtmatr, widths = analyze_circadian_rhythm(os.path.dirname(csv_path))

        


    # --- Summary Sheet (from individual feature CSVs) ---
    if not df_summary.empty:
        summary_ws = wb.create_sheet(title="Summary_Metrics")
        summary_ws.append(df_summary.columns.tolist())
        for r_idx, row in df_summary.iterrows():
            summary_ws.append(row.tolist())

        # Auto-size columns for summary sheet
        for col in summary_ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column letter
            for cell in col:
                try: # Necessary to avoid error on non-string values
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            summary_ws.column_dimensions[column].width = adjusted_width

    # --- Daily Path Lengths Sheet (from individual feature CSVs) ---
    if not df_all_daily_paths.empty:
        daily_ws = wb.create_sheet(title="AllDailyPathLengths")
        daily_ws.append(df_all_daily_paths.columns.tolist())
        for r_idx, row in df_all_daily_paths.iterrows():
            daily_ws.append(row.tolist())
        
        # Auto-size columns
        for col in daily_ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            daily_ws.column_dimensions[column].width = adjusted_width

    # --- NEW: Aggregated Plant Motion Sheet ---
    if aggregated_motion_csv_path and os.path.isfile(aggregated_motion_csv_path):
        try:
            df_aggregated_plant_motion = pd.read_csv(aggregated_motion_csv_path)
            if not df_aggregated_plant_motion.empty:
                agg_ws = wb.create_sheet(title="Aggregated_Plant_Motion")
                agg_ws.append(df_aggregated_plant_motion.columns.tolist())
                for r_idx, row in df_aggregated_plant_motion.iterrows():
                    agg_ws.append(row.tolist())
                
                # Auto-size columns
                for col in agg_ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    agg_ws.column_dimensions[column].width = adjusted_width
                print(f"EXPORT: Added 'Aggregated_Plant_Motion' sheet from {aggregated_motion_csv_path}")
            else:
                print(f"EXPORT WARNING: Aggregated motion CSV is empty: {aggregated_motion_csv_path}")
        except Exception as e:
            print(f"EXPORT ERROR: Could not read or add aggregated motion CSV {aggregated_motion_csv_path}: {e}")


    # Save the Excel workbook
    try:
        wb.save(excel_output_path)
        print(f"EXPORT: Data exported to Excel: {excel_output_path}")
    except Exception as e:
        print(f"EXPORT ERROR: Failed to save Excel workbook to {excel_output_path}: {e}")


# --- Main Execution Block for Testing ---
if __name__ == "__main__":
    # Example usage for testing (uncomment to run standalone)
    # This block assumes you have a directory with 'feature_*.csv' files generated by TRACKMOTION
    # and a mock `overall_datetimes_agg` if you want to test `calculate_and_export_aggregated_motion`
    
    # Example 1: Test calculate_and_export_aggregated_motion
    # mock_motion_features_dir = "C:/Path/To/Your/Test/Plant1/motion_features_1" # Replace with your test path
    # mock_overall_datetimes_agg = [
    #     "01/01/2023 10:00:00", "01/01/2023 10:01:00", "01/01/2023 10:02:00",
    #     "01/01/2023 10:03:00", "01/01/2023 10:04:00", "01/01/2023 10:05:00"
    #
    # if os.path.exists(mock_motion_features_dir):
    #     aggregated_csv = calculate_and_export_aggregated_motion(
    #         mock_motion_features_dir,
    #         mock_overall_datetimes_agg
    #     )
    #     print(f"Aggregated CSV path: {aggregated_csv}")
    #
    #     # Example 2: Test export_to_excel (batch mode)
    #     export_to_excel(
    #         batch_mode=True,
    #         csv_dir=mock_motion_features_dir,
    #         output_filename="batch_motion_report.xlsx",
    #         aggregated_motion_csv_path=aggregated_csv # Pass the generated aggregated CSV
    #     )
    # else:
    #     print(f"Test directory not found: {mock_motion_features_dir}")

    pass